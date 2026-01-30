from __future__ import annotations

import logging
import os
import re
import tkinter as tk
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Tuple
from tkinter import filedialog, messagebox

# Bibliotecas externas
import pdfplumber
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# ==========================================
# 1. CONFIGURAÇÃO DO "MOTOR" DE LEITURA (TESSERACT)
# ==========================================

# Define o caminho fixo onde o Tesseract está instalado no seu computador.
# É necessário apontar para a pasta correta para o Python saber onde procurar.
PASTA_INSTALACAO = r'C:\Users\jose.demorais\AppData\Local\Programs\Tesseract-OCR'

# Cria o caminho completo adicionando o nome do executável (.exe) ao final.
CAMINHO_EXECUTAVEL = os.path.join(PASTA_INSTALACAO, 'tesseract.exe')

# Diz à biblioteca pytesseract: "O motor está aqui, usa este ficheiro".
pytesseract.pytesseract.tesseract_cmd = CAMINHO_EXECUTAVEL

# Cria uma "bandeira" de segurança. Começa como Falso para evitar erros.
OCR_ATIVADO = False

# Verifica se o ficheiro realmente existe antes de tentar usá-lo.
if os.path.exists(CAMINHO_EXECUTAVEL):
    print(f"✅ SUCESSO! Tesseract encontrado em: {CAMINHO_EXECUTAVEL}")
    OCR_ATIVADO = True # Ativa a funcionalidade de leitura de imagens
else:
    # Se não encontrar, avisa no terminal, mas o código continua (só não vai ler imagens).
    print(f"❌ ERRO CRÍTICO: O arquivo não está em: {CAMINHO_EXECUTAVEL}")

# Configurações gerais de visualização de logs e nome da janela
APP_TITLE = "ConciliaPDF — Versão Final (Com Totais)"
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(message)s")

# ==========================================
# 2. ESTRUTURA DE DADOS (A "FICHA" DO FICHEIRO)
# ==========================================
@dataclass(frozen=True)
class PdfItem:
    """
    Isto define o 'molde' de cada linha do Excel.
    Para cada PDF lido, o sistema preenche uma ficha destas.
    """
    file_name: str  # Nome do ficheiro (ex: fatura.pdf)
    file_path: str  # Onde está guardado no computador
    category: str   # Categoria (Receita ou Despesa)
    amount: float   # O valor numérico para cálculo (ex: 1050.50)
    status: str     # O resultado (OK, REVISAR, ERRO)
    method: str     # Explicação de como o valor foi encontrado

# ==========================================
# 3. FUNÇÕES MATEMÁTICAS E DE LIMPEZA
# ==========================================

def br_money_to_float(raw: str) -> float:
    """
    Traduz o formato brasileiro (1.000,00) para formato de computador (1000.00).
    Sem isto, não conseguimos somar os valores.
    """
    if not clean : return 0.0
    
    # Remove tudo o que não for número, vírgula ou ponto (tira letras, R$, espaços)
    clean = re.sub(r"[^\d,\.]", "", str(raw)) #remove tudo o que não for número, vírgula ou ponto (tira letras, R$, espaços)
    if not clean: return 0.0
    
    # Troca a pontuação:
    # 1. Remove o ponto de milhar (1.000 vira 1000)
    # 2. Troca a vírgula decimal por ponto (50,90 vira 50.90)
    clean = clean.replace(".", "").replace(",", ".")
    
    try:
        return float(clean) # Converte texto para número real
    except ValueError:
        return 0.0

def format_br(value: float) -> str:
    """
    Faz o oposto da função anterior.
    Pega no número de cálculo e deixa bonito para o Excel (R$ 1.200,50).
    """
    # Usa um truque com 'X' para trocar ponto por vírgula sem baralhar os dois
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def clean_ocr_text(text: str) -> str:
    """
    Corrige erros visuais do Tesseract.
    Como é uma leitura por imagem, ele às vezes confunde formas parecidas.
    """
    if not text: return ""
    
    # Corrige confusões comuns:
    # '|' vira nada, '!' vira 1, 'l' (ele) vira 1
    t = text.replace("|", "").replace("!", "1").replace("l", "1")
    
    # Corrige símbolos matemáticos colados
    t = t.replace("$=", " ").replace("=", " = ")
    return t
# ==========================================
# LEITURA DO PDF
# ==========================================
def ler_conteudo_pdf(pdf_path: Path) -> Tuple[str, str]:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            paginas_texto = []
            for p in pdf.pages:
                paginas_texto.append(p.extract_text() or "")
            texto_digital = "\n".join(paginas_texto)

            if len(texto_digital.strip()) > 50:
                return texto_digital, "TEXTO_DIGITAL"

            if not OCR_ATIVADO:
                return "", "FALHA: É imagem e Tesseract não foi achado."

            imagem = pdf.pages[0].to_image(resolution=300).original 
            texto_lido = pytesseract.image_to_string(imagem, lang="por")
            return texto_lido, "OCR (IA Visual)"
            
    except Exception as e:
        return "", f"ERRO LEITURA: {str(e)}"
# ==========================================
# EXTRAÇÃO DE VALORES (MODO SENSÍVEL)
# ==========================================
def extrair_valor(text: str) -> Tuple[float, str]:
    # 1. LIMPEZA INICIAL
    # Remove sujeira do OCR (ex: troca '!' por '1')
    text_clean = clean_ocr_text(text)
    
    # Transforma tudo em MAIÚSCULAS. Assim, tanto faz se está escrito "Nota" ou "NOTA".
    text_upper = text_clean.upper()#transforma tudo em maiúsculas
    
    # 2. O GATILHO DO "MODO SENSÍVEL" (A grande mudança)
    # O robô verifica se o texto tem palavras de documentos oficiais importantes.
    # Se encontrar qualquer uma destas, ele muda o comportamento para ser mais preciso.
    eh_documento_oficial = "NOTA DE DÉBITO" in text_upper or "PENALIDADE" in text_upper or "NOTA FISCAL" in text_upper

    # 3. A REDE DE PESCA (REGEX)
    # Esta linha procura TODOS os padrões numéricos que parecem dinheiro brasileiro.
    # Ex: pega "1.000,00", pega "37,88", pega "2025,00".
    todos_valores = re.findall(r"(\d{1,3}(?:\.\d{3})*,\d{2})", text_clean)
    
    lista_floats = []#lista de valores em formato float
    
    # 4. A FILTRAGEM INTELIGENTE
    if todos_valores:
        for v in todos_valores:
            # Traduz o texto para número de computador (troca vírgula por ponto)
            f = br_money_to_float(v)
            
            # --- FILTRO A: BLOQUEIO DE DATAS ---
            # Se o número for exatamente um ano atual ou próximo, IGNORA.
            # Isso impede que a data "25/12/2025" seja lida como R$ 2.025,00.
            if f in [2024.0, 2025.0, 2026.0, 2027.0]:
                continue
            
            # --- FILTRO B: A LÓGICA DE GARANHUNS ---
            if eh_documento_oficial:
                # Se sabemos que é uma Nota/Penalidade (Modo Sensível), confiamos no documento.
                # Aceitamos qualquer valor maior que ZERO (resolve o caso dos R$ 37,88).
                if f > 0: lista_floats.append(f)
            else:
                # Se NÃO sabemos o que é o documento, somos desconfiados.
                # Só aceitamos valores acima de 50 para não pegar número de página ou lixo.
                if f > 50: lista_floats.append(f)

    # 5. A DECISÃO FINAL
    if lista_floats:
        # Se sobraram números válidos, pega o MAIOR de todos.
        # Em 99% das faturas, o maior valor presente na folha é o "Total a Pagar".
        maior_valor = max(lista_floats)
        
        # Define a mensagem de status baseada no modo usado
        msg = "Maior Valor (Modo Sensível)" if eh_documento_oficial else "Maior Valor (>50)"
        return maior_valor, msg

    # Se não achou nada ou tudo foi filtrado, retorna zero.
    return 0.0, "Valor não identificado"

# ==========================================
# GERADOR DE EXCEL (ALTERADO)
# ==========================================
def salvar_excel(caminho: Path, itens: List[PdfItem]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    
    # Cabeçalho
    cabecalho = ["Arquivo", "Categoria", "Valor", "Status", "Método", "Caminho"]
    ws.append(cabecalho)
    
    total_receita = 0.0
    total_despesa = 0.0
    
    # Preenche dados e calcula totais
    for i in itens:
        ws.append([
            i.file_name, i.category, f"R$ {format_br(i.amount)}",
            i.status, i.method, i.file_path
        ])
        
        # Só soma se o status for OK para garantir precisão
        if i.status == "OK":
            if i.category == "Receita":
                total_receita += i.amount
            elif i.category == "Despesa":
                total_despesa += i.amount
    
    # Formatação do Cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="003366")
    
    # --- ÁREA DE TOTAIS (NOVO) ---
    ws.append([]) # Linha em branco
    ws.append([]) # Linha em branco
    
    linha_inicio = ws.max_row + 1
    
    # Adiciona os totais
    ws.append(["RESUMO FINANCEIRO", "", "", "", "", ""])
    ws.append(["(+) TOTAL RECEITAS", "", f"R$ {format_br(total_receita)}", "", "", ""])
    ws.append(["(-) TOTAL DESPESAS", "", f"R$ {format_br(total_despesa)}", "", "", ""])
    
    saldo = total_receita - total_despesa
    ws.append(["(=) SALDO FINAL", "", f"R$ {format_br(saldo)}", "", "", ""])
    
    # Formatação dos Totais
    # Estilo Receita (Verde)
    cell_rec_lbl = ws.cell(row=linha_inicio+1, column=1)
    cell_rec_val = ws.cell(row=linha_inicio+1, column=3)
    cell_rec_lbl.font = Font(bold=True, color="006600")
    cell_rec_val.font = Font(bold=True, color="006600")

    # Estilo Despesa (Vermelho)
    cell_desp_lbl = ws.cell(row=linha_inicio+2, column=1)
    cell_desp_val = ws.cell(row=linha_inicio+2, column=3)
    cell_desp_lbl.font = Font(bold=True, color="CC0000")
    cell_desp_val.font = Font(bold=True, color="CC0000")

    # Estilo Saldo (Azul ou Preto)
    cell_saldo_lbl = ws.cell(row=linha_inicio+3, column=1)
    cell_saldo_val = ws.cell(row=linha_inicio+3, column=3)
    cell_saldo_lbl.font = Font(bold=True, size=12)
    cell_saldo_val.font = Font(bold=True, size=12)
    
    # Ajuste de larguras
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["E"].width = 40
    
    wb.save(caminho)

# ==========================================
# TELA
# ==========================================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("650x500")
        self.path_rec = None
        self.path_desp = None

        tk.Label(self, text="Conciliador Final (Totais e Saldo)", font=("Arial", 14, "bold"), fg="#333").pack(pady=15)
        
        if OCR_ATIVADO:
            status_ocr = "✅ TESSERACT CONECTADO!"
            cor_ocr = "green"
        else:
            status_ocr = "❌ ERRO NO TESSERACT"
            cor_ocr = "red"

        tk.Label(self, text=status_ocr, font=("Arial", 12, "bold"), fg=cor_ocr).pack()

        frame = tk.Frame(self)
        frame.pack(pady=20)

        tk.Button(frame, text="1. Pasta RECEITAS", width=25, height=2, command=self.sel_rec).grid(row=0, column=0, padx=5, pady=5)
        self.lbl_rec = tk.Label(frame, text="Nenhuma pasta selecionada")
        self.lbl_rec.grid(row=0, column=1, sticky="w")

        tk.Button(frame, text="2. Pasta DESPESAS", width=25, height=2, command=self.sel_desp).grid(row=1, column=0, padx=5, pady=5)
        self.lbl_desp = tk.Label(frame, text="Nenhuma pasta selecionada")
        self.lbl_desp.grid(row=1, column=1, sticky="w")

        btn_run = tk.Button(self, text="PROCESSAR E CALCULAR", font=("Arial", 12, "bold"), bg="#008CBA", fg="white", width=30, height=2, command=self.run)
        btn_run.pack(pady=20)

        self.status = tk.Label(self, text="Aguardando...", fg="blue")
        self.status.pack()

    def sel_rec(self):
        p = filedialog.askdirectory()
        if p: 
            self.path_rec = Path(p)
            qtd = len(list(self.path_rec.rglob("*.pdf")))
            self.lbl_rec.config(text=f"{qtd} arquivos encontrados", fg="green")

    def sel_desp(self):
        p = filedialog.askdirectory()
        if p: 
            self.path_desp = Path(p)
            qtd = len(list(self.path_desp.rglob("*.pdf")))
            self.lbl_desp.config(text=f"{qtd} arquivos encontrados", fg="green")

    def run(self):
        if not self.path_rec and not self.path_desp:
            return messagebox.showwarning("Ops", "Selecione as pastas primeiro!")
        
        destino = filedialog.askdirectory(title="Onde salvar o Relatório Final?")
        if not destino: return

        self.status.config(text="Lendo arquivos... (OCR pode demorar)")
        self.update()

        arquivos_rec = list(self.path_rec.rglob("*.pdf")) if self.path_rec else []
        arquivos_desp = list(self.path_desp.rglob("*.pdf")) if self.path_desp else []

        itens = processar_lista(arquivos_rec, "Receita") + processar_lista(arquivos_desp, "Despesa")

        timestamp = datetime.now().strftime("%H%M%S")
        caminho_excel = Path(destino) / f"Relatorio_Final_{timestamp}.xlsx"
        salvar_excel(caminho_excel, itens)

        tot_rec = sum(i.amount for i in itens if i.category == "Receita" and i.status == "OK")
        tot_desp = sum(i.amount for i in itens if i.category == "Despesa" and i.status == "OK")
        saldo = tot_rec - tot_desp
        pendencias = sum(1 for i in itens if i.status != "OK")

        msg = (f"Cálculo Finalizado!\n\n"
               f"(+) Receitas: {format_br(tot_rec)}\n"
               f"(-) Despesas: {format_br(tot_desp)}\n"
               f"(=) SALDO: {format_br(saldo)}\n\n"
               f"Pendências: {pendencias}\n"
               f"Abra o Excel para ver o detalhe.")
        
        messagebox.showinfo("Sucesso", msg)
        self.status.config(text="Processo finalizado.")

if __name__ == "__main__":
    App().mainloop()